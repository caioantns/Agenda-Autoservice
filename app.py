"""Agenda Autoservice - Versão Web

Roda com: python app.py (ambiente local)
Em produção (Railway/Render): gunicorn app:app
"""

import os
import io
from datetime import datetime, date, timedelta
from functools import wraps

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from models import (
    db, Servico, ListaValor, Equipamento, Usuario, STATUS_OPCOES, HORARIOS_OPCOES,
    PERIODO_OPCOES, STATUS_ESTOQUE_OPCOES, codigos_instalados_lista, codigos_instalados_texto,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL", f"sqlite:///{os.path.join(BASE_DIR, 'servicos.db')}"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.secret_key = os.environ.get("SECRET_KEY", "troque-esta-chave-em-producao")

# Usados apenas para criar a primeira conta admin automaticamente, na primeira
# vez que o app roda (quando ainda não existe nenhum usuário cadastrado).
# Depois disso, a gestão de contas é feita pela aba "Usuários" (perfil admin).
ADMIN_INICIAL_USUARIO = os.environ.get("ADMIN_USUARIO", "admin")
ADMIN_INICIAL_SENHA = os.environ.get("ADMIN_SENHA", "servico123")

MESES_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

db.init_app(app)


# ----------------------------------------------------------------------
# Autenticação: contas individuais com perfil admin ou técnico
# ----------------------------------------------------------------------

def login_obrigatorio(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapper


def admin_obrigatorio(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            return jsonify({"erro": "Apenas administradores podem fazer isso."}), 403
        return view(*args, **kwargs)
    return wrapper


def usuario_atual_e_admin():
    return session.get("role") == "admin"


def tecnico_da_sessao():
    """Nome do técnico logado, ou None se for admin (admin não tem filtro)."""
    if session.get("role") == "admin":
        return None
    return session.get("tecnico_nome") or ""


def pode_acessar_servico(servico):
    """Admin acessa tudo. Técnico só acessa serviços atribuídos a ele mesmo."""
    tecnico_sessao = tecnico_da_sessao()
    if tecnico_sessao is None:
        return True
    return servico.tecnico == tecnico_sessao


def resolver_filtro_tecnico():
    """Para admin, respeita o filtro escolhido na tela (querystring). Para
    técnico, ignora o que veio da tela e trava sempre no próprio nome."""
    tecnico_sessao = tecnico_da_sessao()
    if tecnico_sessao is not None:
        return tecnico_sessao
    return request.args.get("tecnico", "Todos")


@app.route("/login", methods=["GET", "POST"])
def login():
    erro = None
    if request.method == "POST":
        usuario_texto = (request.form.get("usuario") or "").strip()
        senha_texto = request.form.get("senha") or ""
        usuario = Usuario.query.filter_by(usuario=usuario_texto, ativo=True).first()
        if usuario and check_password_hash(usuario.senha_hash, senha_texto):
            session["user_id"] = usuario.id
            session["nome"] = usuario.nome
            session["role"] = usuario.role
            session["tecnico_nome"] = usuario.tecnico_nome or ""
            session.permanent = True
            return redirect(url_for("home"))
        erro = "Usuário ou senha incorretos."
    return render_template("login.html", erro=erro)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ----------------------------------------------------------------------
# Utilitários
# ----------------------------------------------------------------------

def validar_data(texto):
    try:
        return datetime.strptime(texto, "%d/%m/%Y").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return None


def validar_valor(texto):
    if texto is None or str(texto).strip() == "":
        return 0.0
    texto = str(texto).strip().replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        valor = float(texto)
        return valor if valor >= 0 else None
    except ValueError:
        return None


def obter_listas():
    listas = {"empresa": [], "tipo_servico": [], "tecnico": []}
    for item in ListaValor.query.order_by(ListaValor.valor).all():
        if item.categoria in listas:
            listas[item.categoria].append(item.valor)
    return listas


def registrar_valor_se_novo(categoria, valor):
    valor = (valor or "").strip()
    if not valor:
        return
    existente = ListaValor.query.filter_by(categoria=categoria, valor=valor).first()
    if not existente:
        db.session.add(ListaValor(categoria=categoria, valor=valor))
        db.session.commit()


def valor_existe_na_lista(categoria, valor):
    """Confirma que um valor (empresa/tipo de serviço/técnico) já está entre os
    pré-definidos. Evita que digitações diferentes da mesma coisa (ex: 'Caio'
    vs 'CAIO') entrem no sistema e quebrem os filtros por técnico."""
    valor = (valor or "").strip()
    if not valor:
        return True  # campo vazio é tratado como "não informado", não é erro aqui
    return ListaValor.query.filter_by(categoria=categoria, valor=valor).first() is not None


def validar_campos_pre_definidos(tipo_servico, empresa, tecnico):
    if not valor_existe_na_lista("tipo_servico", tipo_servico):
        return f'"{tipo_servico}" não está na lista de tipos de serviço cadastrados.'
    if not valor_existe_na_lista("empresa", empresa):
        return f'"{empresa}" não está na lista de empresas cadastradas.'
    if not valor_existe_na_lista("tecnico", tecnico):
        return f'"{tecnico}" não está na lista de técnicos cadastrados.'
    return None


def normalizar_texto(texto):
    """Remove acentos e caixa alta/baixa para comparar tipo de serviço sem depender de digitação exata."""
    import unicodedata
    texto = (texto or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn")


def tipo_exige_retirado(tipo_servico):
    return normalizar_texto(tipo_servico) == "retirada"


def tipo_exige_instalado(tipo_servico):
    return normalizar_texto(tipo_servico) == "instalacao"


def tipo_e_manutencao(tipo_servico):
    return normalizar_texto(tipo_servico) == "manutencao"


def validar_equipamentos_instalados_existem(empresa, codigos_instalados):
    for codigo_instalado in codigos_instalados:
        existe = Equipamento.query.filter_by(empresa=empresa, codigo=codigo_instalado).first()
        if not existe:
            return (
                f'O equipamento "{codigo_instalado}" não está cadastrado no estoque dessa empresa. '
                f"Cadastre-o na aba Estoque antes de instalar."
            )
    return None


def validar_equipamento_obrigatorio(
    status, tipo_servico, empresa, equipamento_retirado_codigo,
    equipamento_instalado_codigo, houve_troca_equipamento,
):
    """Ao marcar como Feito, exige:
    - Equipamento Retirado, se o tipo for Retirada.
    - Ao menos 1 Equipamento Instalado, se o tipo for Instalação (pode ser mais de
      um) — e, nesse caso, cada código precisa já existir cadastrado no estoque
      da empresa.
    - Se o tipo for Manutenção, primeiro pergunta se houve troca de equipamento:
      se sim, exige os dois (retirado + ao menos 1 instalado, nas mesmas regras
      acima); se não, não exige nenhum código.
    `equipamento_instalado_codigo` chega aqui já como texto com os códigos
    separados por vírgula."""
    if status != "Feito":
        return None

    if tipo_exige_retirado(tipo_servico):
        if not (equipamento_retirado_codigo or "").strip():
            return "Para marcar como Feito, informe o ID do equipamento retirado."
        return None

    if tipo_exige_instalado(tipo_servico):
        codigos_instalados = codigos_instalados_lista(equipamento_instalado_codigo)
        if not codigos_instalados:
            return "Para marcar como Feito, informe o ID do equipamento instalado."
        return validar_equipamentos_instalados_existem(empresa, codigos_instalados)

    if tipo_e_manutencao(tipo_servico):
        if houve_troca_equipamento is None:
            return "Antes de concluir, informe se houve troca de equipamento nessa manutenção."
        if houve_troca_equipamento is True:
            if not (equipamento_retirado_codigo or "").strip():
                return "Para marcar como Feito, informe o ID do equipamento retirado."
            codigos_instalados = codigos_instalados_lista(equipamento_instalado_codigo)
            if not codigos_instalados:
                return "Para marcar como Feito, informe o ID do equipamento instalado."
            return validar_equipamentos_instalados_existem(empresa, codigos_instalados)

    return None


def montar_codigos_instalados(dados):
    """Lê os códigos de equipamento instalado enviados pelo front. Aceita a lista
    nova (`equipamento_instalado_codigos`, um item por campo adicionado na tela)
    e, por compatibilidade, também o campo antigo de um único código."""
    lista = dados.get("equipamento_instalado_codigos")
    if lista is None:
        unico = (dados.get("equipamento_instalado_codigo") or "").strip()
        lista = [unico] if unico else []
    return codigos_instalados_texto(lista)


def normalizar_houve_troca(valor):
    """Aceita True/False vindos do JSON; qualquer outra coisa (ausente, null,
    string) vira None, tratado como 'ainda não respondido'."""
    if isinstance(valor, bool):
        return valor
    return None


def registrar_equipamento_retirado_no_estoque(empresa, codigo):
    """Quando um equipamento é retirado de um cliente, ele volta para o estoque da
    empresa marcado como Indisponível (pendente de inspeção). Se ainda não estiver
    cadastrado no estoque dessa empresa, cria o registro automaticamente."""
    codigo = (codigo or "").strip()
    if not codigo or not empresa:
        return
    existente = Equipamento.query.filter_by(empresa=empresa, codigo=codigo).first()
    if existente:
        existente.status = "Indisponível"
    else:
        db.session.add(Equipamento(empresa=empresa, modelo="", codigo=codigo, status="Indisponível"))
    db.session.commit()


def marcar_equipamento_instalado_utilizado(empresa, equipamento_instalado_codigo):
    """Quando equipamento(s) são instalados num cliente, cada um passa a Utilizado.
    `equipamento_instalado_codigo` chega aqui como texto com os códigos separados
    por vírgula (pode ser um só)."""
    if not empresa:
        return
    for codigo in codigos_instalados_lista(equipamento_instalado_codigo):
        existente = Equipamento.query.filter_by(empresa=empresa, codigo=codigo).first()
        if existente:
            existente.status = "Utilizado"
    db.session.commit()


def verificar_conflito(tecnico, data, horario, servico_id_ignorar=None):
    if not tecnico or not horario:
        return None
    query = Servico.query.filter(
        Servico.tecnico == tecnico,
        Servico.data == data,
        Servico.horario == horario,
        Servico.status != "Cancelado",
    )
    if servico_id_ignorar is not None:
        query = query.filter(Servico.id != servico_id_ignorar)
    return query.first()


# ----------------------------------------------------------------------
# Páginas
# ----------------------------------------------------------------------

@app.route("/")
@login_obrigatorio
def index():
    return redirect(url_for("home"))


@app.route("/home")
@login_obrigatorio
def home():
    return render_template("home.html", aba_ativa="home", role=session.get("role"), nome=session.get("nome", ""))


@app.route("/api/home/stats")
@login_obrigatorio
def home_stats():
    tecnico_filtro = tecnico_da_sessao()

    hoje_str = date.today().strftime("%d/%m/%Y")
    query_hoje = Servico.query.filter_by(data=hoje_str)
    if tecnico_filtro is not None:
        query_hoje = query_hoje.filter_by(tecnico=tecnico_filtro)
    servicos_hoje = query_hoje.count()

    hoje = date.today()
    limite = hoje + timedelta(days=6)
    proximos_7_dias = 0
    todos_servicos = Servico.query.filter(Servico.status != "Cancelado")
    if tecnico_filtro is not None:
        todos_servicos = todos_servicos.filter(Servico.tecnico == tecnico_filtro)
    for s in todos_servicos.all():
        try:
            d = datetime.strptime(s.data, "%d/%m/%Y").date()
        except ValueError:
            continue
        if hoje <= d <= limite:
            proximos_7_dias += 1

    total_empresas = ListaValor.query.filter_by(categoria="empresa").count()
    total_equipamentos_indisponiveis = 0
    if usuario_atual_e_admin():
        for eq in Equipamento.query.all():
            situacao, _ = buscar_uso_equipamento(eq.empresa, eq.codigo)
            if situacao != "Disponível":
                total_equipamentos_indisponiveis += 1

    return jsonify({
        "servicos_hoje": servicos_hoje,
        "proximos_7_dias": proximos_7_dias,
        "total_empresas": total_empresas,
        "equipamentos_fora_do_estoque": total_equipamentos_indisponiveis,
    })


@app.route("/cadastro")
@login_obrigatorio
def cadastro():
    listas = obter_listas()
    hoje = date.today().strftime("%d/%m/%Y")
    return render_template(
        "cadastro.html", listas=listas, status_opcoes=STATUS_OPCOES,
        horarios_opcoes=HORARIOS_OPCOES, hoje=hoje, aba_ativa="cadastro",
        role=session.get("role"), tecnico_nome=session.get("tecnico_nome", ""),
    )


@app.route("/agenda")
@login_obrigatorio
def agenda():
    listas = obter_listas()
    hoje = date.today().strftime("%d/%m/%Y")
    return render_template(
        "agenda.html", tecnicos=listas["tecnico"], listas=listas, status_opcoes=STATUS_OPCOES,
        horarios_opcoes=HORARIOS_OPCOES, hoje=hoje, aba_ativa="agenda",
        role=session.get("role"), tecnico_nome=session.get("tecnico_nome", ""),
    )


@app.route("/resumo")
@login_obrigatorio
def resumo():
    listas = obter_listas()
    hoje = date.today()
    return render_template(
        "resumo.html", tecnicos=listas["tecnico"], meses=MESES_PT,
        mes_atual=hoje.month, ano_atual=hoje.year, aba_ativa="resumo",
        role=session.get("role"), tecnico_nome=session.get("tecnico_nome", ""),
    )


@app.route("/estoque")
@admin_obrigatorio
def estoque():
    listas = obter_listas()
    return render_template(
        "estoque.html", empresas=listas["empresa"], aba_ativa="estoque",
        role=session.get("role"),
    )


@app.route("/usuarios")
@admin_obrigatorio
def usuarios_pagina():
    listas = obter_listas()
    return render_template("usuarios.html", aba_ativa="usuarios", tecnicos=listas["tecnico"])


# ----------------------------------------------------------------------
# API: Serviços (usada pelo JavaScript das páginas)
# ----------------------------------------------------------------------

@app.route("/api/servicos", methods=["POST"])
@login_obrigatorio
def criar_servico():
    dados = request.get_json(force=True)

    data_valida = validar_data(dados.get("data"))
    if not data_valida:
        return jsonify({"erro": "Data inválida. Use o formato dd/mm/aaaa."}), 400

    placa = (dados.get("placa") or "").strip().upper()

    tipo_servico = (dados.get("tipo_servico") or "").strip()
    if not tipo_servico:
        return jsonify({"erro": "Tipo de Serviço é obrigatório."}), 400

    empresa = (dados.get("empresa") or "").strip()
    if not empresa:
        return jsonify({"erro": "Empresa é obrigatória."}), 400

    valor = validar_valor(dados.get("valor"))
    if valor is None:
        return jsonify({"erro": "Valor inválido."}), 400

    horario = (dados.get("horario") or "").strip()
    tecnico_sessao = tecnico_da_sessao()
    tecnico = tecnico_sessao if tecnico_sessao is not None else (dados.get("tecnico") or "").strip()
    status = dados.get("status") if dados.get("status") in STATUS_OPCOES else STATUS_OPCOES[0]

    erro_pre_definido = validar_campos_pre_definidos(tipo_servico, empresa, tecnico)
    if erro_pre_definido:
        return jsonify({"erro": erro_pre_definido}), 400

    equipamento_retirado_codigo = (dados.get("equipamento_retirado_codigo") or "").strip()
    equipamento_instalado_codigo = montar_codigos_instalados(dados)
    houve_troca_equipamento = normalizar_houve_troca(dados.get("houve_troca_equipamento"))
    if not tipo_e_manutencao(tipo_servico):
        houve_troca_equipamento = None

    erro_equipamento = validar_equipamento_obrigatorio(
        status, tipo_servico, empresa, equipamento_retirado_codigo,
        equipamento_instalado_codigo, houve_troca_equipamento,
    )
    if erro_equipamento:
        return jsonify({"erro": erro_equipamento}), 400

    conflito = verificar_conflito(tecnico, data_valida, horario)
    if conflito and not dados.get("ignorar_conflito"):
        return jsonify({
            "conflito": True,
            "mensagem": f"{tecnico} já tem um serviço em {data_valida} às {horario} "
                        f"(Empresa: {conflito.empresa}). Deseja salvar mesmo assim?",
        }), 409

    servico = Servico(
        data=data_valida, horario=horario, placa=placa, tipo_servico=tipo_servico,
        empresa=empresa, endereco=(dados.get("endereco") or "").strip(),
        tecnico=tecnico, status=status, valor=valor,
        observacoes=(dados.get("observacoes") or "").strip(),
        equipamento_retirado_codigo=equipamento_retirado_codigo,
        equipamento_instalado_codigo=equipamento_instalado_codigo,
        houve_troca_equipamento=houve_troca_equipamento,
    )
    db.session.add(servico)
    db.session.commit()

    if equipamento_retirado_codigo:
        registrar_equipamento_retirado_no_estoque(empresa, equipamento_retirado_codigo)
    if equipamento_instalado_codigo:
        marcar_equipamento_instalado_utilizado(empresa, equipamento_instalado_codigo)

    registrar_valor_se_novo("empresa", empresa)
    registrar_valor_se_novo("tipo_servico", tipo_servico)
    registrar_valor_se_novo("tecnico", tecnico)

    return jsonify({"ok": True, "servico": servico.to_dict()})


@app.route("/api/servicos/<int:servico_id>", methods=["PUT"])
@login_obrigatorio
def atualizar_servico(servico_id):
    servico = Servico.query.get_or_404(servico_id)
    if not pode_acessar_servico(servico):
        return jsonify({"erro": "Você não tem permissão para editar este serviço."}), 403

    dados = request.get_json(force=True)

    data_valida = validar_data(dados.get("data"))
    if not data_valida:
        return jsonify({"erro": "Data inválida. Use o formato dd/mm/aaaa."}), 400

    placa = (dados.get("placa") or "").strip().upper()

    tipo_servico = (dados.get("tipo_servico") or "").strip()
    empresa = (dados.get("empresa") or "").strip()
    if not tipo_servico or not empresa:
        return jsonify({"erro": "Tipo de Serviço e Empresa são obrigatórios."}), 400

    valor = validar_valor(dados.get("valor"))
    if valor is None:
        return jsonify({"erro": "Valor inválido."}), 400

    periodo = dados.get("periodo") if dados.get("periodo") in PERIODO_OPCOES else ""
    if not periodo:
        return jsonify({"erro": "Selecione o período (Manhã, Integral ou Tarde)."}), 400
    tecnico_sessao = tecnico_da_sessao()
    tecnico = tecnico_sessao if tecnico_sessao is not None else (dados.get("tecnico") or "").strip()
    status = dados.get("status") if dados.get("status") in STATUS_OPCOES else servico.status


    erro_pre_definido = validar_campos_pre_definidos(tipo_servico, empresa, tecnico)
    if erro_pre_definido:
        return jsonify({"erro": erro_pre_definido}), 400

    equipamento_retirado_codigo = (dados.get("equipamento_retirado_codigo") or "").strip()
    equipamento_instalado_codigo = montar_codigos_instalados(dados)
    houve_troca_equipamento = normalizar_houve_troca(dados.get("houve_troca_equipamento"))
    if not tipo_e_manutencao(tipo_servico):
        houve_troca_equipamento = None

    erro_equipamento = validar_equipamento_obrigatorio(
        status, tipo_servico, empresa, equipamento_retirado_codigo,
        equipamento_instalado_codigo, houve_troca_equipamento,
    )
    if erro_equipamento:
        return jsonify({"erro": erro_equipamento}), 400

    conflito = verificar_conflito(tecnico, data_valida, periodo, servico_id_ignorar=servico.id)
    if conflito and not dados.get("ignorar_conflito"):
        return jsonify({
            "conflito": True,
            "mensagem": f"{tecnico} já tem um serviço em {data_valida} às {horario} "
                        f"(Empresa: {conflito.empresa}). Deseja salvar mesmo assim?",
        }), 409

    servico.data = data_valida
    servico.periodo = periodo
    servico.placa = placa
    servico.tipo_servico = tipo_servico
    servico.empresa = empresa
    servico.endereco = (dados.get("endereco") or "").strip()
    servico.tecnico = tecnico
    servico.status = status
    servico.valor = valor
    servico.equipamento_retirado_codigo = equipamento_retirado_codigo
    servico.equipamento_instalado_codigo = equipamento_instalado_codigo
    servico.houve_troca_equipamento = houve_troca_equipamento
    if "observacoes" in dados:
        servico.observacoes = (dados.get("observacoes") or "").strip()
    db.session.commit()

    if equipamento_retirado_codigo:
        registrar_equipamento_retirado_no_estoque(empresa, equipamento_retirado_codigo)
    if equipamento_instalado_codigo:
        marcar_equipamento_instalado_utilizado(empresa, equipamento_instalado_codigo)

    registrar_valor_se_novo("empresa", empresa)
    registrar_valor_se_novo("tipo_servico", tipo_servico)
    registrar_valor_se_novo("tecnico", tecnico)

    return jsonify({"ok": True, "servico": servico.to_dict()})


@app.route("/api/servicos/<int:servico_id>/status", methods=["PATCH"])
@login_obrigatorio
def atualizar_status(servico_id):
    """Atalho rápido usado no celular: mudar só o status (Feito/Em andamento/Cancelado)."""
    servico = Servico.query.get_or_404(servico_id)
    if not pode_acessar_servico(servico):
        return jsonify({"erro": "Você não tem permissão para alterar este serviço."}), 403

    dados = request.get_json(force=True)
    novo_status = dados.get("status")
    if novo_status not in STATUS_OPCOES:
        return jsonify({"erro": "Status inválido."}), 400

    erro_equipamento = validar_equipamento_obrigatorio(
        novo_status, servico.tipo_servico, servico.empresa,
        servico.equipamento_retirado_codigo, servico.equipamento_instalado_codigo,
        servico.houve_troca_equipamento,
    )
    if erro_equipamento:
        return jsonify({"erro": erro_equipamento, "precisa_equipamento": True}), 400

    servico.status = novo_status
    db.session.commit()

    if novo_status == "Feito":
        if servico.equipamento_retirado_codigo:
            registrar_equipamento_retirado_no_estoque(servico.empresa, servico.equipamento_retirado_codigo)
        if servico.equipamento_instalado_codigo:
            marcar_equipamento_instalado_utilizado(servico.empresa, servico.equipamento_instalado_codigo)

    return jsonify({"ok": True, "servico": servico.to_dict()})


@app.route("/api/servicos/<int:servico_id>/observacoes", methods=["PATCH"])
@login_obrigatorio
def atualizar_observacoes(servico_id):
    servico = Servico.query.get_or_404(servico_id)
    if not pode_acessar_servico(servico):
        return jsonify({"erro": "Você não tem permissão para alterar este serviço."}), 403

    dados = request.get_json(force=True)
    servico.observacoes = (dados.get("observacoes") or "").strip()
    db.session.commit()
    return jsonify({"ok": True, "servico": servico.to_dict()})


@app.route("/api/servicos/<int:servico_id>", methods=["DELETE"])
@login_obrigatorio
def excluir_servico(servico_id):
    servico = Servico.query.get_or_404(servico_id)
    if not pode_acessar_servico(servico):
        return jsonify({"erro": "Você não tem permissão para excluir este serviço."}), 403

    db.session.delete(servico)
    db.session.commit()
    return jsonify({"ok": True})


@app.route("/api/servicos/<int:servico_id>/duplicar", methods=["POST"])
@login_obrigatorio
def duplicar_servico(servico_id):
    original = Servico.query.get_or_404(servico_id)
    if not pode_acessar_servico(original):
        return jsonify({"erro": "Você não tem permissão para duplicar este serviço."}), 403

    dados = request.get_json(force=True) or {}
    tecnico_sessao = tecnico_da_sessao()

    novo = Servico(
        data=validar_data(dados.get("data")) or original.data,
        horario=dados.get("horario", original.horario),
        placa=original.placa,
        tipo_servico=original.tipo_servico,
        empresa=original.empresa,
        endereco=original.endereco,
        tecnico=tecnico_sessao if tecnico_sessao is not None else original.tecnico,
        status="Agendado",
        valor=original.valor,
        observacoes="",
    )
    db.session.add(novo)
    db.session.commit()
    return jsonify({"ok": True, "servico": novo.to_dict()})


@app.route("/api/agenda/<data_str>")
@login_obrigatorio
def servicos_do_dia(data_str):
    """data_str no formato dd-mm-aaaa (compatível com URL)."""
    try:
        data_fmt = datetime.strptime(data_str, "%d-%m-%Y").strftime("%d/%m/%Y")
    except ValueError:
        return jsonify({"erro": "Data inválida."}), 400

    tecnico_filtro = resolver_filtro_tecnico()
    query = Servico.query.filter_by(data=data_fmt)
    if tecnico_filtro not in ("Todos", ""):
        query = query.filter_by(tecnico=tecnico_filtro)

    ORDEM_PERIODO = {"Manhã": 0, "Integral": 1, "Tarde": 2}
    servicos = query.all()
    servicos.sort(key=lambda s: ORDEM_PERIODO.get(s.periodo, 9))
    return jsonify([s.to_dict() for s in servicos])


@app.route("/api/agenda/mes/<int:ano>/<int:mes>")
@login_obrigatorio
def resumo_do_mes(ano, mes):
    """Retorna, por dia do mês, quantos serviços há e se estão todos feitos/pendentes/cancelados."""
    tecnico_filtro = resolver_filtro_tecnico()
    # SQLite não indexa bem strings dd/mm/aaaa para extrair mês/ano, então filtramos em Python.
    todos = Servico.query.all()
    if tecnico_filtro not in ("Todos", ""):
        todos = [s for s in todos if s.tecnico == tecnico_filtro]

    por_dia = {}
    for s in todos:
        try:
            d = datetime.strptime(s.data, "%d/%m/%Y").date()
        except ValueError:
            continue
        if d.year != ano or d.month != mes:
            continue
        por_dia.setdefault(d.day, []).append(s.status)

    resultado = {}
    for dia, status_list in por_dia.items():
        if any(st in ("Agendado", "Em andamento") for st in status_list):
            situacao = "pendente"
        elif all(st == "Cancelado" for st in status_list):
            situacao = "cancelado"
        elif all(st in ("Frustrado", "Cancelado") for st in status_list):
            situacao = "frustrado"
        else:
            situacao = "feito"
        resultado[dia] = {"quantidade": len(status_list), "situacao": situacao}

    return jsonify(resultado)


@app.route("/api/proximos-7-dias")
@login_obrigatorio
def proximos_7_dias():
    try:
        tecnico = request.args.get("tecnico", "").strip()
        hoje = date.today()
        limite = hoje + timedelta(days=7)

        # Consulta base
        query = Servico.query

        # Filtra técnico apenas se for um nome real (ignora "Todos" ou vazio)
        if tecnico and tecnico.lower() != "todos":
            query = query.filter(Servico.tecnico == tecnico)

        # Busca todos e filtra o intervalo de forma segura
        todos_servicos = query.all()
        resultado = []

        for s in todos_servicos:
            # Garante a leitura da data como string YYYY-MM-DD
            data_str = str(s.data).split(" ")[0] if s.data else ""

            # Converte com segurança para comparar
            try:
                data_obj = date.fromisoformat(data_str)
                if hoje <= data_obj <= limite:
                    # Serialização segura
                    dados = (
                        s.to_dict()
                        if hasattr(s, "to_dict")
                        else {
                            "id": s.id,
                            "data": data_str,
                            "horario": getattr(s, "horario", ""),
                            "periodo": getattr(s, "periodo", ""),
                            "tipo_servico": getattr(s, "tipo_servico", ""),
                            "empresa": getattr(s, "empresa", ""),
                            "endereco": getattr(s, "endereco", ""),
                            "tecnico": getattr(s, "tecnico", ""),
                            "status": getattr(s, "status", "Agendado"),
                            "valor": getattr(s, "valor", 0.0),
                            "equipamento_instalado": getattr(
                                s, "equipamento_instalado", None
                            ),
                            "observacoes": getattr(s, "observacoes", ""),
                        }
                    )
                    resultado.append(dados)
            except Exception:
                continue

        # Ordena por data e horário
        resultado.sort(key=lambda x: (x.get("data", ""), x.get("horario", "")))
        return jsonify(resultado)

    except Exception as e:
        print(f"Erro em proximos_7_dias: {e}")
        return jsonify({"erro": str(e)}), 500


@app.route("/api/servicos/exportar")
@login_obrigatorio
def exportar_servicos_excel():
    """Baixa todos os serviços visíveis para o usuário logado num arquivo
    .xlsx — serve como backup manual e para uso fora do sistema."""
    tecnico_filtro = tecnico_da_sessao()
    query = Servico.query.order_by(Servico.id)
    if tecnico_filtro is not None:
        query = query.filter_by(tecnico=tecnico_filtro)
    servicos = query.all()

    colunas = [
        "ID", "Data", "Período", "Horário", "Placa", "Tipo de Serviço", "Empresa", "Endereço",
        "Técnico", "Status", "Valor", "Equipamento Retirado", "Equipamento Instalado",
        "Observações",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Serviços"
    ws.append(colunas)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="2F5496")

    for s in servicos:
        ws.append([
            s.id, s.data, s.periodo or "", s.horario or "", s.placa or "", s.tipo_servico, s.empresa,
            s.endereco or "", s.tecnico or "", s.status, s.valor or 0.0,
            s.equipamento_retirado_codigo or "", s.equipamento_instalado_codigo or "",
            s.observacoes or "",
        ])

    for i, coluna in enumerate(colunas, start=1):
        largura = max(len(coluna), 12) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = largura

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"backup_servicos_{datetime.now().strftime('%Y-%m-%d_%Hh%M')}.xlsx"
    return send_file(
        buffer, as_attachment=True, download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/servicos/importar", methods=["POST"])
@admin_obrigatorio
def importar_servicos_excel():
    """Restaura serviços a partir de um .xlsx gerado pelo próprio botão de
    exportar. ADICIONA os serviços do arquivo (não apaga nem substitui o que
    já existe) — pensado para restaurar dados após trocar de banco de dados."""
    arquivo = request.files.get("arquivo")
    if not arquivo:
        return jsonify({"erro": "Nenhum arquivo enviado."}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(arquivo, data_only=True)
        ws = wb.active
    except Exception:
        return jsonify({"erro": "Não consegui abrir esse arquivo. Confirme que é um .xlsx válido."}), 400

    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return jsonify({"erro": "Planilha vazia."}), 400

    cabecalho = [str(c or "").strip() for c in linhas[0]]
    esperado = [
        "ID", "Data", "Período", "Horário", "Placa", "Tipo de Serviço", "Empresa", "Endereço",
        "Técnico", "Status", "Valor", "Equipamento Retirado", "Equipamento Instalado",
        "Observações",
    ]
    if cabecalho != esperado:
        return jsonify({
            "erro": "As colunas desse arquivo não batem com o formato esperado "
                    "(use um arquivo gerado pelo botão 'Exportar tudo para Excel')."
        }), 400

    importados = 0
    ignorados = 0
    for linha in linhas[1:]:
        if not linha or not any(linha):
            continue
        (_, data_str, periodo, horario, placa, tipo_servico, empresa, endereco,
         tecnico, status, valor, eq_retirado, eq_instalado, observacoes) = (list(linha) + [None] * 14)[:14]

        data_valida = validar_data(str(data_str or "").strip())
        tipo_servico = str(tipo_servico or "").strip()
        empresa = str(empresa or "").strip()
        if not data_valida or not tipo_servico or not empresa:
            ignorados += 1
            continue

        valor_final = validar_valor(valor)
        if valor_final is None:
            valor_final = 0.0

        status_final = str(status or "").strip()
        if status_final not in STATUS_OPCOES:
            status_final = "Agendado"

        servico = Servico(
            data=data_valida,
            periodo=str(periodo or "").strip() if str(periodo or "").strip() in PERIODO_OPCOES else "",
            horario=str(horario or "").strip(),
            placa=str(placa or "").strip().upper(),
            tipo_servico=tipo_servico,
            empresa=empresa,
            endereco=str(endereco or "").strip(),
            tecnico=str(tecnico or "").strip(),
            status=status_final,
            valor=valor_final,
            equipamento_retirado_codigo=str(eq_retirado or "").strip(),
            equipamento_instalado_codigo=str(eq_instalado or "").strip(),
            observacoes=str(observacoes or "").strip(),
        )
        db.session.add(servico)
        importados += 1

        registrar_valor_se_novo("empresa", servico.empresa)
        registrar_valor_se_novo("tipo_servico", servico.tipo_servico)
        if servico.tecnico:
            registrar_valor_se_novo("tecnico", servico.tecnico)

    db.session.commit()

    return jsonify({"ok": True, "importados": importados, "ignorados": ignorados})


@app.route("/api/resumo/<int:ano>/<int:mes>")
@login_obrigatorio
def resumo_mensal(ano, mes):
    tecnico_sessao = tecnico_da_sessao()
    if tecnico_sessao is not None:
        tecnicos = [tecnico_sessao]
    else:
        tecnicos = [t for (t,) in db.session.query(ListaValor.valor).filter_by(categoria="tecnico").all()]

    todos = Servico.query.filter_by(status="Feito").all()
    if tecnico_sessao is not None:
        todos = [s for s in todos if s.tecnico == tecnico_sessao]

    contagem = {t: {"feitos": 0, "valor": 0.0} for t in tecnicos}
    for s in todos:
        try:
            d = datetime.strptime(s.data, "%d/%m/%Y").date()
        except ValueError:
            continue
        if d.year != ano or d.month != mes:
            continue
        tecnico = s.tecnico or "(sem técnico)"
        item = contagem.setdefault(tecnico, {"feitos": 0, "valor": 0.0})
        item["feitos"] += 1
        item["valor"] += s.valor or 0.0

    total_feitos = sum(v["feitos"] for v in contagem.values())
    total_valor = sum(v["valor"] for v in contagem.values())

    return jsonify({
        "por_tecnico": [{"tecnico": t, **v} for t, v in sorted(contagem.items())],
        "total_feitos": total_feitos,
        "total_valor": total_valor,
    })


# ----------------------------------------------------------------------
# API: Listas gerenciáveis (empresas, tipos de serviço, técnicos)
# ----------------------------------------------------------------------

@app.route("/api/listas/<categoria>", methods=["GET"])
@login_obrigatorio
def listar_valores(categoria):
    itens = ListaValor.query.filter_by(categoria=categoria).order_by(ListaValor.valor).all()
    return jsonify([i.valor for i in itens])


@app.route("/api/listas/<categoria>", methods=["POST"])
@login_obrigatorio
def adicionar_valor(categoria):
    if categoria == "tecnico" and not usuario_atual_e_admin():
        return jsonify({"erro": "Apenas administradores podem gerenciar a lista de técnicos."}), 403
    dados = request.get_json(force=True)
    valor = (dados.get("valor") or "").strip()
    if not valor:
        return jsonify({"erro": "Valor não pode ser vazio."}), 400
    registrar_valor_se_novo(categoria, valor)
    return jsonify({"ok": True})


@app.route("/api/listas/<categoria>/<valor>", methods=["DELETE"])
@login_obrigatorio
def remover_valor(categoria, valor):
    if categoria == "tecnico" and not usuario_atual_e_admin():
        return jsonify({"erro": "Apenas administradores podem gerenciar a lista de técnicos."}), 403
    item = ListaValor.query.filter_by(categoria=categoria, valor=valor).first()
    if item:
        db.session.delete(item)
        db.session.commit()
    return jsonify({"ok": True})


# ----------------------------------------------------------------------
# API: Estoque de equipamentos por empresa
# ----------------------------------------------------------------------

def buscar_ultima_os_do_equipamento(empresa, codigo):
    """Só para exibir informação (não decide mais o status) — retorna o serviço
    mais recente (não cancelado) que usou esse código, seja como retirado ou como
    um dos instalados (uma OS pode ter mais de um equipamento instalado)."""
    candidatos = (
        Servico.query.filter(Servico.empresa == empresa, Servico.status != "Cancelado")
        .order_by(Servico.id.desc()).all()
    )
    for servico in candidatos:
        if servico.equipamento_retirado_codigo == codigo:
            return servico
        if codigo in codigos_instalados_lista(servico.equipamento_instalado_codigo):
            return servico
    return None


@app.route("/api/estoque/<empresa>")
@login_obrigatorio
def listar_estoque(empresa):
    itens = Equipamento.query.filter_by(empresa=empresa).order_by(Equipamento.codigo).all()
    resultado = []
    for item in itens:
        dados = item.to_dict()
        dados["disponivel"] = item.status == "Disponível"  # mantém compatibilidade com os formulários
        servico_vinculado = buscar_ultima_os_do_equipamento(empresa, item.codigo)
        dados["os_vinculada"] = None
        if servico_vinculado:
            dados["os_vinculada"] = {
                "id": servico_vinculado.id,
                "data": servico_vinculado.data,
                "periodo": servico_vinculado.periodo,
                "tipo_servico": servico_vinculado.tipo_servico,
                "tecnico": servico_vinculado.tecnico,
                "status": servico_vinculado.status,
            }
        resultado.append(dados)
    return jsonify(resultado)


@app.route("/api/estoque/<int:item_id>/status", methods=["PATCH"])
@admin_obrigatorio
def mudar_status_equipamento(item_id):
    item = Equipamento.query.get_or_404(item_id)
    dados = request.get_json(force=True)
    novo_status = dados.get("status")
    if novo_status not in STATUS_ESTOQUE_OPCOES:
        return jsonify({"erro": "Status inválido."}), 400
    item.status = novo_status
    db.session.commit()
    return jsonify({"ok": True, "item": item.to_dict()})


@app.route("/api/estoque", methods=["POST"])
@admin_obrigatorio
def criar_equipamento():
    dados = request.get_json(force=True)
    empresa = (dados.get("empresa") or "").strip()
    codigo = (dados.get("codigo") or "").strip()
    if not empresa or not codigo:
        return jsonify({"erro": "Empresa e ID do equipamento são obrigatórios."}), 400

    existente = Equipamento.query.filter_by(empresa=empresa, codigo=codigo).first()
    if existente:
        return jsonify({"erro": f'Já existe um equipamento com o ID "{codigo}" cadastrado nessa empresa.'}), 400

    item = Equipamento(empresa=empresa, modelo=(dados.get("modelo") or "").strip(), codigo=codigo)
    db.session.add(item)
    db.session.commit()
    return jsonify({"ok": True, "item": item.to_dict()})


@app.route("/api/estoque/<int:item_id>", methods=["PUT"])
@admin_obrigatorio
def atualizar_equipamento(item_id):
    item = Equipamento.query.get_or_404(item_id)
    dados = request.get_json(force=True)

    codigo = (dados.get("codigo") or "").strip()
    if not codigo:
        return jsonify({"erro": "ID do equipamento é obrigatório."}), 400

    existente = Equipamento.query.filter(
        Equipamento.empresa == item.empresa, Equipamento.codigo == codigo, Equipamento.id != item.id
    ).first()
    if existente:
        return jsonify({"erro": f'Já existe outro equipamento com o ID "{codigo}" cadastrado nessa empresa.'}), 400

    item.modelo = (dados.get("modelo") or "").strip()
    item.codigo = codigo
    db.session.commit()
    return jsonify({"ok": True, "item": item.to_dict()})


@app.route("/api/estoque/<int:item_id>", methods=["DELETE"])
@admin_obrigatorio
def excluir_equipamento(item_id):
    item = Equipamento.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return jsonify({"ok": True})


# ----------------------------------------------------------------------
# API: Usuários (contas de acesso) - apenas administradores
# ----------------------------------------------------------------------

@app.route("/api/usuarios")
@admin_obrigatorio
def listar_usuarios():
    usuarios = Usuario.query.order_by(Usuario.nome).all()
    return jsonify([u.to_dict() for u in usuarios])


@app.route("/api/usuarios", methods=["POST"])
@admin_obrigatorio
def criar_usuario():
    dados = request.get_json(force=True)
    nome = (dados.get("nome") or "").strip()
    usuario_login = (dados.get("usuario") or "").strip().lower()
    senha = dados.get("senha") or ""
    role = dados.get("role") if dados.get("role") in ("admin", "tecnico") else "tecnico"
    tecnico_nome = (dados.get("tecnico_nome") or "").strip()

    if not nome or not usuario_login or not senha:
        return jsonify({"erro": "Nome, usuário e senha são obrigatórios."}), 400
    if len(senha) < 4:
        return jsonify({"erro": "A senha precisa ter pelo menos 4 caracteres."}), 400
    if role == "tecnico" and not tecnico_nome:
        return jsonify({"erro": "Informe qual Técnico (da lista) essa conta representa."}), 400

    if Usuario.query.filter_by(usuario=usuario_login).first():
        return jsonify({"erro": f'Já existe uma conta com o usuário "{usuario_login}".'}), 400

    novo = Usuario(
        nome=nome, usuario=usuario_login, senha_hash=generate_password_hash(senha),
        role=role, tecnico_nome=tecnico_nome if role == "tecnico" else "",
    )
    db.session.add(novo)
    db.session.commit()

    if role == "tecnico":
        registrar_valor_se_novo("tecnico", tecnico_nome)

    return jsonify({"ok": True, "usuario": novo.to_dict()})


@app.route("/api/usuarios/<int:usuario_id>", methods=["PUT"])
@admin_obrigatorio
def atualizar_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    dados = request.get_json(force=True)

    nome = (dados.get("nome") or "").strip()
    role = dados.get("role") if dados.get("role") in ("admin", "tecnico") else usuario.role
    tecnico_nome = (dados.get("tecnico_nome") or "").strip()

    if not nome:
        return jsonify({"erro": "Nome é obrigatório."}), 400
    if role == "tecnico" and not tecnico_nome:
        return jsonify({"erro": "Informe qual Técnico (da lista) essa conta representa."}), 400

    usuario.nome = nome
    usuario.role = role
    usuario.tecnico_nome = tecnico_nome if role == "tecnico" else ""
    usuario.ativo = bool(dados.get("ativo", usuario.ativo))

    nova_senha = dados.get("senha") or ""
    if nova_senha:
        if len(nova_senha) < 4:
            return jsonify({"erro": "A senha precisa ter pelo menos 4 caracteres."}), 400
        usuario.senha_hash = generate_password_hash(nova_senha)

    db.session.commit()

    if role == "tecnico" and tecnico_nome:
        registrar_valor_se_novo("tecnico", tecnico_nome)

    return jsonify({"ok": True, "usuario": usuario.to_dict()})


@app.route("/api/usuarios/<int:usuario_id>", methods=["DELETE"])
@admin_obrigatorio
def excluir_usuario(usuario_id):
    if usuario_id == session.get("user_id"):
        return jsonify({"erro": "Você não pode excluir a própria conta enquanto está logado nela."}), 400
    usuario = Usuario.query.get_or_404(usuario_id)
    db.session.delete(usuario)
    db.session.commit()
    return jsonify({"ok": True})


# ----------------------------------------------------------------------
# Inicialização
# ----------------------------------------------------------------------

with app.app_context():
    db.create_all()
    if Usuario.query.count() == 0:
        admin_inicial = Usuario(
            nome="Administrador", usuario=ADMIN_INICIAL_USUARIO,
            senha_hash=generate_password_hash(ADMIN_INICIAL_SENHA), role="admin",
        )
        db.session.add(admin_inicial)
        db.session.commit()
        print(
            f"Conta admin inicial criada: usuário '{ADMIN_INICIAL_USUARIO}'. "
            f"Entre e crie as demais contas na aba Usuários."
        )


if __name__ == "__main__":
    porta = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=porta, debug=True)
